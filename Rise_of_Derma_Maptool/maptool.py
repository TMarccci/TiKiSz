import openpyxl

file = "./MatrixPlanner.xlsx"
wb = openpyxl.load_workbook(file, read_only=True)
ws = wb.active

with open("./Finalized.txt", "w", encoding="utf-8") as f:
    for i in range(1, ws.max_row+1):
        for j in range(1, ws.max_column+1):
            if ws.cell(row=i, column=j).value is not None:
                cellvalue = ws.cell(row=i, column=j).value
                numbers = [str(x+1) for x in range(9)]
                
                if cellvalue == "@":
                    f.write(f"{j-2};{i-2};WALL\n")
                elif cellvalue == "S":
                    f.write(f"{j-2};{i-2};PLAYERSPAWN\n")
                elif cellvalue == "C":
                    f.write(f"{j-2};{i-2};FINISHLEVEL\n")
                elif cellvalue == "K":
                    f.write(f"{j-2};{i-2};CRYSTAL\n")
                elif cellvalue in numbers:
                    f.write(f"{j-2};{i-2};ENEMY;{cellvalue}\n")
    
    f.close()
            